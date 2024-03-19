import os
import glob
import asyncio
import aiohttp
import openpyxl
from aiofile import async_open


async def extract_text_from_pdf(pdf_path):
    if not os.path.exists(pdf_path):
        print(f"PDF file does not exist: {pdf_path}")
        return ""
    try:
        async with async_open(pdf_path, "rb") as file:
            doc = await fitz.open(stream=await file.read(), filetype="pdf")
            text = ""
            for page in doc:
                text += page.get_text()
            return text
    except Exception as e:
        print(f"Error extracting text from PDF: {e}")
        return ""


async def read_bibtex_file(file_path):
    try:
        async with async_open(
            file_path, "r", encoding="utf-8", errors="replace"
        ) as bibtex_file:
            bibtex_text = await bibtex_file.read()
            return bibtex_text
    except Exception as e:
        print(f"Error reading file {file_path}: {e}")
        return None


async def download_and_process_paper(doi, bibtex_dir, pdf_dir, sheet, row_num):
    bibtex_file_name = f"{doi.replace('/', '_')}.bib"
    bibtex_file_path = os.path.join(bibtex_dir, bibtex_file_name)
    pdf_file_path = os.path.join(pdf_dir, f"{doi.replace('/', '_')}.pdf")

    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(
                f"https://api.crossref.org/works/{doi}/transform/application/x-bibtex"
            ) as response:
                bibtex_text = await response.text()
                async with async_open(
                    bibtex_file_path, "w", encoding="utf-8"
                ) as bibtex_file:
                    await bibtex_file.write(bibtex_text)
    except Exception as e:
        print(f"Failed to download BibTeX for DOI: {doi}. Error: {e}")
        bibtex_text = ""

    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(
                f"https://doi.org/{doi}", allow_redirects=True
            ) as response:
                async with async_open(pdf_file_path, "wb") as pdf_file:
                    async for chunk in response.content.iter_chunked(1024):
                        await pdf_file.write(chunk)
    except Exception as e:
        print(f"Failed to download PDF for DOI: {doi}. Error: {e}")
        pdf_file_path = "PDF not available"

    full_text = (
        await extract_text_from_pdf(pdf_file_path)
        if os.path.exists(pdf_file_path)
        else ""
    )

    sheet.cell(row=row_num, column=sheet.max_column + 1, value=bibtex_text)
    sheet.cell(row=row_num, column=sheet.max_column, value=full_text)
    sheet.cell(
        row=row_num,
        column=sheet.max_column,
        value=pdf_file_path if os.path.exists(pdf_file_path) else "PDF not available",
    )


async def update_xlsx_with_bibtex_and_pdf_text(xlsx_path, semaphore):
    base_dir = os.path.dirname(xlsx_path)
    bibtex_dir = os.path.join(base_dir, "bibtex")
    pdf_dir = os.path.join(base_dir, "pdfs")
    os.makedirs(bibtex_dir, exist_ok=True)
    os.makedirs(pdf_dir, exist_ok=True)

    workbook = openpyxl.load_workbook(xlsx_path)
    sheet = workbook.active

    tasks = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        doi = row[sheet.cell(row=1, column=sheet.max_column).value]
        row_num = row[0].row
        task = asyncio.create_task(
            download_and_process_paper(doi, bibtex_dir, pdf_dir, sheet, row_num)
        )
        tasks.append(task)

        # Limit the number of concurrent tasks to 5
        if len(tasks) >= 5:
            await asyncio.gather(*tasks)
            tasks = []

    # Process any remaining tasks
    if tasks:
        await asyncio.gather(*tasks)

    # Ensure each subsection has at least 4 entries
    subsections = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        subsection = row[1]  # Assuming the subsection is in the second column
        if subsection not in subsections:
            subsections[subsection] = []
        subsections[subsection].append(row)

    for subsection, entries in subsections.items():
        while len(entries) < 4:
            empty_entry = [None] * (
                sheet.max_column - 1
            )  # Create an empty entry with None values
            entries.append(empty_entry)
            sheet.append(empty_entry)

    workbook.save(xlsx_path)


async def process_all_xlsx_files(directory):
    semaphore = asyncio.Semaphore(5)  # Limit the number of concurrent tasks to 5
    tasks = []

    for xlsx_path in glob.glob(os.path.join(directory, "**", "*.xlsx"), recursive=True):
        print(f"Processing {xlsx_path}...")
        task = asyncio.create_task(
            update_xlsx_with_bibtex_and_pdf_text(xlsx_path, semaphore)
        )
        tasks.append(task)

    await asyncio.gather(*tasks)


if __name__ == "__main__":
    directory = "searches"
    asyncio.run(process_all_xlsx_files(directory))
