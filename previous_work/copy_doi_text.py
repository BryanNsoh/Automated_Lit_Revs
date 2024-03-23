import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import logging
import time
import os

# Configure logging
logging.basicConfig(
    filename="doi_scraper.log",
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s",
)


def open_xlsx_file(file_path, sheet_name):
    logging.debug(f"Attempting to open XLSX file: {file_path}, sheet: {sheet_name}")
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        logging.debug(
            f"Successfully opened XLSX file: {file_path}, sheet: {sheet_name}"
        )
        return workbook, sheet
    except FileNotFoundError:
        logging.error(f"File not found: {file_path}")
        raise
    except KeyError:
        logging.error(f"Sheet not found: {sheet_name}")
        raise
    except Exception as e:
        logging.error(f"Error opening XLSX file: {str(e)}")
        raise


def extract_doi_links_and_identifiers(sheet, doi_column, identifier_column):
    logging.debug(
        f"Extracting DOI links and identifiers from sheet: {sheet.title}, DOI column: {doi_column}, identifier column: {identifier_column}"
    )
    doi_links_and_identifiers = []
    for row in sheet.iter_rows(min_row=2):
        doi_cell = next(
            (
                cell
                for cell in row
                if sheet.cell(row=1, column=cell.column).value == doi_column
            ),
            None,
        )
        identifier_cell = next(
            (
                cell
                for cell in row
                if sheet.cell(row=1, column=cell.column).value == identifier_column
            ),
            None,
        )
        if doi_cell and identifier_cell:
            doi_link = doi_cell.value
            identifier = identifier_cell.value
            if doi_link and identifier:
                # Validate DOI link format (simplified example)
                if doi_link.startswith("https://doi.org/"):
                    doi_links_and_identifiers.append((doi_link, identifier))
                    logging.debug(
                        f"Valid DOI link found: {doi_link}, identifier: {identifier}"
                    )
                else:
                    logging.warning(f"Invalid DOI link: {doi_link}")
            else:
                logging.debug(
                    f"Empty or None DOI link or identifier found in row {doi_cell.row}"
                )
    logging.debug(
        f"Extracted {len(doi_links_and_identifiers)} DOI links and identifiers from sheet: {sheet.title}"
    )
    return doi_links_and_identifiers


def initialize_webdriver():
    logging.debug("Initializing WebDriver")
    try:
        options = webdriver.ChromeOptions()
        options.add_argument("--headless")  # Run in headless mode
        driver = webdriver.Chrome(options=options)
        logging.debug("WebDriver initialized successfully")
        return driver
    except Exception as e:
        logging.error(f"Error initializing WebDriver: {str(e)}")
        raise


def navigate_to_doi_link(driver, doi_link):
    logging.debug(f"Navigating to DOI link: {doi_link}")
    try:
        driver.get(doi_link)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        logging.debug(f"Successfully navigated to DOI link: {doi_link}")
    except TimeoutException:
        logging.warning(f"Timeout navigating to: {doi_link}")
    except Exception as e:
        logging.error(f"Error navigating to: {doi_link} - {str(e)}")


def extract_text_content(driver):
    logging.debug("Extracting text content from page")
    try:
        body = driver.find_element(By.TAG_NAME, "body")
        text_content = body.text
        logging.debug(
            f"Extracted text content: {text_content[:100]}..."
        )  # Log first 100 characters
        return text_content
    except NoSuchElementException:
        logging.warning("Body element not found.")
    except Exception as e:
        logging.error(f"Error extracting text content: {str(e)}")
    return ""


def remove_special_characters(text):
    logging.debug(
        f"Removing special characters from text: {text[:100]}..."
    )  # Log first 100 characters
    cleaned_text = "".join(c for c in text if c.isprintable())
    logging.debug(f"Cleaned text: {cleaned_text[:100]}...")  # Log first 100 characters
    return cleaned_text


def paste_content_to_xlsx(sheet, identifier, paste_column, content):
    logging.debug(
        f"Pasting content to XLSX sheet: {sheet.title}, identifier: {identifier}, column: {paste_column}"
    )
    try:
        for row in sheet.iter_rows(min_row=2):
            identifier_cell = next(
                (
                    cell
                    for cell in row
                    if sheet.cell(row=1, column=cell.column).value
                    == "unique_identifier"
                ),
                None,
            )
            if identifier_cell and identifier_cell.value == identifier:
                paste_cell = next(
                    (
                        cell
                        for cell in row
                        if sheet.cell(row=1, column=cell.column).value == paste_column
                    ),
                    None,
                )
                if paste_cell and not paste_cell.value:
                    paste_cell.value = remove_special_characters(content)
                    logging.debug(f"Content pasted to cell: {paste_cell.coordinate}")
                    return
        logging.warning(
            f"No matching identifier found for {identifier} or paste cell already has content"
        )
    except Exception as e:
        logging.error(f"Error pasting content to XLSX: {str(e)}")


def save_xlsx_file(workbook, file_path):
    logging.debug(f"Saving XLSX file: {file_path}")
    try:
        workbook.save(file_path)
        logging.info(f"Updated XLSX file saved: {file_path}")
    except Exception as e:
        logging.error(f"Error saving XLSX file: {str(e)}")


def process_sheet(file_path, sheet_name, doi_column, identifier_column, paste_column):
    driver = None
    try:
        logging.info(f"Processing sheet: {sheet_name}")

        # Open the XLSX file
        workbook, sheet = open_xlsx_file(file_path, sheet_name)

        # Extract DOI links and identifiers from the XLSX file
        doi_links_and_identifiers = extract_doi_links_and_identifiers(
            sheet, doi_column, identifier_column
        )
        logging.info(
            f"Extracted {len(doi_links_and_identifiers)} DOI links and identifiers from sheet {sheet_name}."
        )

        # Initialize WebDriver
        driver = initialize_webdriver()

        # Process each DOI link
        for doi_link, identifier in doi_links_and_identifiers:
            logging.info(f"Processing DOI link: {doi_link}, identifier: {identifier}")

            # Navigate to the DOI link
            navigate_to_doi_link(driver, doi_link)

            # Extract the text content
            content = extract_text_content(driver)

            # Paste the content back into the XLSX file
            paste_content_to_xlsx(sheet, identifier, paste_column, content)

            logging.info(
                f"Content pasted for DOI link: {doi_link}, identifier: {identifier}"
            )

            # Add a delay to avoid overwhelming the website (adjust as needed)
            time.sleep(3)

        # Save the updated XLSX file
        save_xlsx_file(workbook, file_path)

    except Exception as e:
        logging.error(
            f"An error occurred while processing sheet {sheet_name}: {str(e)}"
        )

    finally:
        # Close the WebDriver if it was initialized
        if driver is not None:
            driver.quit()
            logging.debug("WebDriver closed")


def main():
    logging.info("Starting DOI Scraper")

    # Set up variables
    file_path = "documents/master.xlsx"
    sheet_names = [str(i) for i in range(1, 9)]  # Process sheets 1 to 8
    doi_column = "doi"
    identifier_column = "unique_identifier"  # Assuming the unique identifier column is named "unique_identifier"
    paste_column = "full_text"

    logging.debug(f"File path: {file_path}")
    logging.debug(f"Sheet names: {sheet_names}")
    logging.debug(f"DOI column: {doi_column}")
    logging.debug(f"Identifier column: {identifier_column}")
    logging.debug(f"Paste column: {paste_column}")

    for sheet_name in sheet_names:
        process_sheet(
            file_path, sheet_name, doi_column, identifier_column, paste_column
        )

    logging.info("DOI Scraper finished")


if __name__ == "__main__":
    main()
