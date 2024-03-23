import os
import glob
import asyncio
import aiohttp
import openpyxl
from aiofile import async_open
import fitz


def remove_special_characters(text):
    return "".join(c for c in text if c.isprintable())


async def extract_text_from_pdf(pdf_path):
    if not os.path.exists(pdf_path):
        print(f"PDF file does not exist: {pdf_path}")
        return ""
    try:
        async with async_open(pdf_path, "rb") as file:
            doc = fitz.open(stream=await file.read(), filetype="pdf")
            text = ""
            for page in doc:
                text += page.get_text()
            return remove_special_characters(text)
    except Exception as e:
        print(f"Error extracting text from PDF: {e}")
        return ""


def download_paper(doi, download_path):
    download_cmd = f'scidownl download --doi {doi} --out {download_path}/{doi.replace("/", "_")}.pdf'
    os.system(download_cmd)


async def download_and_process_paper(doi, pdf_dir, sheet, row):
    pdf_file_path = os.path.join(pdf_dir, f"{doi.replace('/', '_')}.pdf")

    if not os.path.exists(pdf_file_path):
        download_paper(doi, pdf_dir)

    full_text = await extract_text_from_pdf(pdf_file_path)
    sheet.cell(row=row, column=sheet.max_column, value=full_text)


async def update_xlsx_with_pdf_text(xlsx_path, sheet_name, semaphore):
    base_dir = os.path.dirname(xlsx_path)
    pdf_dir = os.path.join(base_dir, "pdfs")
    os.makedirs(pdf_dir, exist_ok=True)

    workbook = openpyxl.load_workbook(xlsx_path)
    try:
        sheet = workbook[sheet_name]
    except KeyError:
        print(f"Sheet '{sheet_name}' not found in {xlsx_path}")
        return

    doi_col = None
    full_text_col = None
    for cell in sheet[1]:
        if cell.value == "doi":
            doi_col = cell.column
        elif cell.value == "full_text":
            full_text_col = cell.column

    if not doi_col:
        print(f"No 'doi' column found in sheet '{sheet_name}' of {xlsx_path}")
        return

    if not full_text_col:
        print(f"No 'full_text' column found in sheet '{sheet_name}' of {xlsx_path}")
        return

    tasks = []
    for row in sheet.iter_rows(min_row=2):
        doi = row[doi_col - 1].value
        full_text = row[full_text_col - 1].value

        if doi and not full_text:
            task = asyncio.create_task(
                download_and_process_paper(doi, pdf_dir, sheet, row[0].row)
            )
            tasks.append(task)

        # Limit the number of concurrent tasks to 5
        if len(tasks) >= 5:
            await asyncio.gather(*tasks)
            tasks = []

    # Process any remaining tasks
    if tasks:
        await asyncio.gather(*tasks)

    workbook.save(xlsx_path)


async def process_xlsx_file(file_path, sheet_names):
    semaphore = asyncio.Semaphore(5)  # Limit the number of concurrent tasks to 5
    tasks = []

    for sheet_name in sheet_names:
        print(f"Processing sheet '{sheet_name}' in {file_path}...")
        task = asyncio.create_task(
            update_xlsx_with_pdf_text(file_path, sheet_name, semaphore)
        )
        tasks.append(task)

    await asyncio.gather(*tasks)


if __name__ == "__main__":
    file_path = "documents/master.xlsx"
    sheet_names = [str(i) for i in range(1, 9)]  # Process sheets 1 to 8

    asyncio.run(process_xlsx_file(file_path, sheet_names))
