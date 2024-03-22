import os
import openpyxl


def update_doi_column(xlsx_path, sheet_name):
    workbook = openpyxl.load_workbook(xlsx_path)

    try:
        sheet = workbook[sheet_name]
    except KeyError:
        print(f"Sheet '{sheet_name}' not found in {xlsx_path}")
        return

    doi_column = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == "doi":
            doi_column = col
            break

    if not doi_column:
        print(f"No 'doi' column found in sheet '{sheet_name}' of {xlsx_path}")
        return

    for row in sheet.iter_rows(min_row=2):
        doi = row[doi_column - 1].value
        if doi:
            if not doi.startswith("https://doi.org/"):
                doi = f"https://doi.org/{doi}"
                row[doi_column - 1].value = doi
            row[doi_column - 1].hyperlink = doi

    workbook.save(xlsx_path)


def process_xlsx_file(file_path, sheet_names):
    for sheet_name in sheet_names:
        print(f"Processing sheet '{sheet_name}' in {file_path}...")
        update_doi_column(file_path, sheet_name)


if __name__ == "__main__":
    file_path = "documents/master.xlsx"
    sheet_names = [str(i) for i in range(1, 9)]  # Process sheets 1 to 8

    process_xlsx_file(file_path, sheet_names)
